from openpyxl import Workbook

CHARACTER_TYPES = ["scoundrel", "brute"]

def createCharacterSheet(characterName, characterClass, characterLevel):

    # Ensure the characterClass the user is trying to make exists
    if not (characterClass.lower() in CHARACTER_TYPES):
        str = " "
        raise Exception("Character Class: " + characterClass + " is not valid. Must be of type: " + str.join(CHARACTER_TYPES))

    # Instantiate character specifics
    character = {
        "name": characterName,
        "class": characterClass,
        "level": characterLevel,
        "health": 0,
        "strength": 0,
        "dexterity": 0
    }

    match character["class"].lower():
        case "brute":
            character["health"] = 12 + (3 * character["level"])
            character["strength"] = 2 + character["level"]

        case "scoundrel":
            character["health"] = 8 + (2 * character["level"])
            character["dexterity"] = 1 + (2 * character["level"])

    return character

# Press the green button in the gutter to run the script.
if __name__ == '__main__':

    wb = Workbook()
    ws = wb.active
    ws.title = "Gloomhaven"

    # set spreadsheet headers
    ws['A1'] = "Name"
    ws['B1'] = "Class"
    ws['C1'] = "Level"
    ws['D1'] = "Health"
    ws['E1'] = "Strength"
    ws['F1'] = "Dexterity"

    characterList = []
    try:

        continueCharacterBuilding = True

        while continueCharacterBuilding:
            try:

                # User inputs
                characterNameString = input("Character name? ")
                characterClassString = input("Character class? ")

                # Check that the Character Class is Valid
                if not (characterClassString.lower() in CHARACTER_TYPES):
                    str = " "
                    raise Exception("Character Class: " + characterClassString + " is not valid. Must be of type: " + str.join(
                        CHARACTER_TYPES))
                    print("")

                characterLevelNum = input("Character level? ")

                # Build Characters
                characterList.append(createCharacterSheet(characterNameString, characterClassString, int(characterLevelNum)))

            except Exception as error:
                print(error)

            # Ask user if they wish to continue
            continueBuildingString = input("Continue creating characters? y/n ")
            print("")
            if not continueBuildingString.lower() == 'y':
                continueCharacterBuilding = False

    except Exception as error:
        print(error)

    rowNumber = 1

    # Start appending Character info to spreadsheet
    for playerCharacter in characterList:
        rowNumber += 1
        # Trait
        ws['A' + rowNumber.__str__()] = playerCharacter["name"]
        ws['B' + rowNumber.__str__()] = playerCharacter["class"]
        ws['C' + rowNumber.__str__()] = playerCharacter["level"]
        ws['D' + rowNumber.__str__()] = playerCharacter["health"]
        ws['E' + rowNumber.__str__()] = playerCharacter["strength"]
        ws['F' + rowNumber.__str__()] = playerCharacter["dexterity"]

    wb.save('/var/www/Character-Sheet.xlsx')
